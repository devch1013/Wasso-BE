from io import BytesIO

import boto3
from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from api.event.models import Attendance, AttendanceStatus, Event
from api.generation.models import Generation, GenMember


def create_attendance_excel(generation: Generation) -> str:
    """
    특정 기수의 출석 정보를 엑셀 파일로 생성하고 S3에 업로드합니다.

    Args:
        generation: Generation 모델

    Returns:
        생성된 엑셀 파일의 S3 URL
    """
    # 색상 정의
    HEADER_FILL = PatternFill(
        start_color="E6E6E6", end_color="E6E6E6", fill_type="solid"
    )
    STATS_FILL = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )
    PRESENT_FILL = PatternFill(
        start_color="D9EBD9", end_color="D9EBD9", fill_type="solid"
    )
    LATE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ABSENT_FILL = PatternFill(
        start_color="FFD9D9", end_color="FFD9D9", fill_type="solid"
    )
    BOLD_FONT = Font(bold=True)

    # 데이터 가져오기
    events = Event.objects.filter(generation=generation).order_by("date", "start_time")
    generation_mappings = GenMember.objects.filter(
        generation=generation
    ).select_related("member__user")
    attendances = Attendance.objects.filter(
        event__in=events, generation_mapping__in=generation_mappings
    )

    # 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = f"{generation.club.name} - {generation.name} - 출석 정보"

    # 헤더 생성
    headers = [f"이름(총 {len(generation_mappings)}명)"]
    for event in events:
        event_header = f"{event.date.strftime('%m/%d')} {event.title}"
        headers.append(event_header)
    headers.extend(["출석", "지각", "결석"])

    # 헤더 입력 및 스타일 적용
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = BOLD_FONT

    # 출석 데이터 입력
    row = 2
    for mapping in generation_mappings:
        present_count = late_count = absent_count = 0
        ws.cell(row=row, column=1, value=mapping.member.user.username)

        for col, event in enumerate(events, 2):
            attendance = (
                attendances.filter(event=event, generation_mapping=mapping)
                .order_by("-modified_at")
                .first()
            )

            status = (
                "출석"
                if attendance and attendance.status == AttendanceStatus.PRESENT
                else "지각"
                if attendance and attendance.status == AttendanceStatus.LATE
                else "결석"
                if attendance and attendance.status == AttendanceStatus.ABSENT
                else "-"
            )

            cell = ws.cell(row=row, column=col, value=status)

            # 상태별 스타일 및 카운트
            if status == "출석":
                cell.fill = PRESENT_FILL
                present_count += 1
            elif status == "지각":
                cell.fill = LATE_FILL
                late_count += 1
            elif status == "결석":
                cell.fill = ABSENT_FILL
                absent_count += 1

        # 통계 추가
        for col, value in enumerate(
            [present_count, late_count, absent_count], len(events) + 2
        ):
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill = STATS_FILL

        row += 1

    # 이벤트별 통계 추가
    stats_labels = ["출석", "지각", "결석"]
    for idx, label in enumerate(stats_labels):
        ws.cell(row=row + idx, column=1, value=label).font = BOLD_FONT

        for col in range(2, len(events) + 2):
            column_data = [ws.cell(row=r, column=col).value for r in range(2, row)]
            count = column_data.count(label)
            cell = ws.cell(row=row + idx, column=col, value=str(count))
            cell.fill = STATS_FILL
            cell.font = BOLD_FONT

    # 열 너비 자동 조정
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # 파일 이름 생성
    filename = f"{generation.club.name}_{generation.name}_출석정보_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # 엑셀 파일을 메모리에 저장
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # S3에 업로드
    s3_client = boto3.client(
        "s3",
        aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
        region_name=settings.AWS_S3_REGION_NAME,
    )

    # S3 경로 설정 (excel 폴더 아래에 저장)
    s3_path = f"excel/{filename}"

    # S3에 업로드
    s3_client.upload_fileobj(
        excel_file,
        settings.AWS_STORAGE_BUCKET_NAME,
        s3_path,
        ExtraArgs={
            "ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        },
    )

    return s3_path
